import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
import threading
import tkinter as tk
from tkinter import filedialog, ttk, Text, Scrollbar, messagebox

cancel_flag = False

def get_decimal_points(number):
    str_num = str(number).split(".")
    return len(str_num[1]) if len(str_num) > 1 else 0

def log_message(txt_log, message):
    txt_log.insert(tk.END, message + "\n")
    txt_log.yview_moveto(1.0)
    root.update_idletasks()

def manipulate_data(df, txt_log):
    if cancel_flag:
        return df

    # Duplicate 25% of random rows
    num_rows_to_duplicate = int(0.25 * len(df))
    duplicated_rows = df.sample(n=num_rows_to_duplicate)
    df = pd.concat([df, duplicated_rows], axis=0).reset_index(drop=True)
    log_message(txt_log, f"Duplicated {num_rows_to_duplicate} rows.")

    # Omit 25% of random rows
    num_rows_to_omit = int(0.25 * len(df))
    drop_indices = np.random.choice(df.index, num_rows_to_omit, replace=False)
    df = df.drop(drop_indices).reset_index(drop=True)
    log_message(txt_log, f"Omitted {num_rows_to_omit} rows.")

    return df

def randomize_data(df, formula_cells, progress_var, txt_log):
    problem_columns = []
    total_cols = len(df.columns)
    for index, col in enumerate(df.columns):
        if cancel_flag:
            return df, problem_columns

        try:
            if df[col].dtype == 'float64' or df[col].dtype == 'int64':
                adjustment = (df[col].max() - df[col].min()) * 0.25
                df[col] = df[col].apply(lambda x: round(x + np.random.uniform(-adjustment, adjustment), get_decimal_points(x)) if (x, col) not in formula_cells else x)
            elif df[col].dtype == 'object':
                df[col] = df[col].sample(frac=1).reset_index(drop=True)
            progress_portion = 100 / total_cols
            current_progress = progress_var.get() + progress_portion
            progress_var.set(current_progress)
            log_message(txt_log, f"Processed column {col}.")
        except Exception as e:
            problem_columns.append((col, str(e)))

    return df, problem_columns

def stop_process():
    global cancel_flag
    cancel_flag = True
    log_message(txt_log, "Cancellation requested...")

def reset_ui():
    btn.config(text="Choose and Anonymize Excel File", command=lambda: threading.Thread(target=choose_and_anonymize).start())
    progress_bar_1.config(mode='determinate')
    progress_bar_2.config(mode='determinate')
    progress_bar_3.config(mode='determinate')
    progress_1.set(0)
    progress_2.set(0)
    progress_3.set(0)

def choose_and_anonymize():
    global cancel_flag
    cancel_flag = False

    try:
        log_message(txt_log, "Starting to process the file...")
        file_path = filedialog.askopenfilename()
        if not file_path:
            return

        btn.config(text="Cancel", command=stop_process)
        progress_bar_1.config(mode='indeterminate')
        progress_bar_1.start(10)
        log_message(txt_log, "Reading Excel file...")
        df = pd.read_excel(file_path)
        log_message(txt_log, "Excel file read successfully.")
        progress_bar_1.stop()
        progress_bar_1.config(mode='determinate')
        progress_1.set(100)

        if cancel_flag:
            log_message(txt_log, "Process cancelled by the user.")
            reset_ui()
            return

        df = manipulate_data(df, txt_log)

        directory, original_name = os.path.split(file_path)
        name, extension = os.path.splitext(original_name)
        default_name = f"{name}_anon{extension}"
        save_path = filedialog.asksaveasfilename(initialdir=directory, initialfile=default_name, defaultextension=".xlsx")
        if not save_path:
            reset_ui()
            return

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        formula_cells = [(cell.value, cell.column_letter) for row in ws.iter_rows() for cell in row if cell.value and isinstance(cell.value, str) and "=" in cell.value]
        wb.close()

        log_message(txt_log, "Preparing for anonymization...")
        progress_2.set(0)

        df_anon, problem_columns = randomize_data(df, formula_cells, progress_2, txt_log)
        progress_2.set(100)

        if cancel_flag:
            log_message(txt_log, "Process cancelled by the user.")
            reset_ui()
            return

        log_message(txt_log, "Starting saving process...")
        progress_bar_3.config(mode='indeterminate')
        progress_bar_3.start(10)
        df_anon.to_excel(save_path, index=False)
        progress_bar_3.stop()
        progress_bar_3.config(mode='determinate')
        progress_3.set(100)
        log_message(txt_log, "Data saved successfully.")

        if problem_columns:
            modal_message = "Finished with errors."
            for col, error_msg in problem_columns:
                log_message(txt_log, f"Column: {col} - Error: {error_msg}")
        else:
            modal_message = "Finished without errors."

        log_message(txt_log, modal_message)
        log_message(txt_log, "=====================================\n\n")
        messagebox.showinfo("Process Completed", modal_message)
    except Exception as e:
        log_message(txt_log, f"Error: {str(e)}")
        messagebox.showerror("Error", str(e))

    reset_ui()

root = tk.Tk()
root.title("Excel Data Anonymizer")

frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

label = ttk.Label(frame, text="Excel Data Anonymizer", font=("Arial", 20))
label.grid(row=0, column=0, pady=20)

btn = ttk.Button(frame, text="Choose and Anonymize Excel File", command=lambda: threading.Thread(target=choose_and_anonymize).start())
btn.grid(row=1, column=0, pady=20)

label_1 = ttk.Label(frame, text="Reading Excel File Progress")
label_1.grid(row=2, column=0, pady=(10, 0))
progress_1 = tk.DoubleVar()
progress_bar_1 = ttk.Progressbar(frame, orient="horizontal", length=300, variable=progress_1)
progress_bar_1.grid(row=3, column=0)

label_2 = ttk.Label(frame, text="Anonymizing Data Progress")
label_2.grid(row=4, column=0, pady=(10, 0))
progress_2 = tk.DoubleVar()
progress_bar_2 = ttk.Progressbar(frame, orient="horizontal", length=300, variable=progress_2)
progress_bar_2.grid(row=5, column=0)

label_3 = ttk.Label(frame, text="Saving Data Progress")
label_3.grid(row=6, column=0, pady=(10, 0))
progress_3 = tk.DoubleVar()
progress_bar_3 = ttk.Progressbar(frame, orient="horizontal", length=300, variable=progress_3)
progress_bar_3.grid(row=7, column=0)

txt_log = Text(frame, wrap=tk.WORD, width=60, height=10)
txt_log.grid(row=8, column=0, pady=10, sticky='w')
scroll = Scrollbar(frame, command=txt_log.yview)
scroll.grid(row=8, column=1, pady=10, sticky='nsew')
txt_log.config(yscrollcommand=scroll.set)

root.mainloop()
